import streamlit as st
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font
from io import BytesIO
import base64

# === Function to convert local image to base64 ===
# def get_base64_image(image_path):
#     with open(image_path, "rb") as image_file:
#         encoded = base64.b64encode(image_file.read()).decode()
#         return f"data:image/png;base64,{encoded}"

# === Column Headers ===
column_headers = [
    "Form No", "Start Date", "Ref No", "Invoice No", "Customer Name", "Email Id", "Mobile No", "Country", "State",
    "City", "Zipcode", "Credit Card Type", "Credit Card Number", "Courier Name", "Sales Date", "Time",
    "Total Amount", "Discount", "Net Amount", "Amount Word", "Product Name", "Rate", "Quantity",
    "Agency Name", "Agency Email Id", "Agency Mobile No", "Dosage", "Blood Group", "Sex", "Medicine",
    "Tablets", "Stm Code", "Stm Name", "Date Of Birth", "Remarks"
]

exclude_headers = {"Form No", "Start Date", "Amount Word", "Remarks"}
filtered_headers = [h for h in column_headers if h not in exclude_headers]

# === Extract field value below label ===
def extract_value(text, field):
    try:
        lines = text.splitlines()
        idx = lines.index(field)
        return lines[idx + 1].strip()
    except:
        return ""

# === Extract all red-colored spans from PDF ===
def extract_red_spans(doc):
    red_texts = []
    for page in doc:
        blocks = page.get_text("dict")["blocks"]
        for block in blocks:
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    if span.get("color") == 16711680:  # Red RGB
                        txt = span["text"].strip()
                        if txt:
                            red_texts.append(txt)
    return red_texts

# === Streamlit UI ===
st.set_page_config(page_title="PDF to Excel", layout="wide")
st.title("📄 PDF → Excel ")
uploaded_files = st.file_uploader("Upload PDF files", type="pdf", accept_multiple_files=True)

if uploaded_files:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    for uploaded_file in uploaded_files:
        if ws.max_row > 1:
            ws.append([])  # empty row between files

        doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
        plain_text = "\n".join(page.get_text() for page in doc)
        red_texts = extract_red_spans(doc)

        # Row 1: Headers
        row_num = ws.max_row + 1
        for col, header in enumerate(filtered_headers, start=1):
            ws.cell(row=row_num, column=col, value=header).font = Font(bold=True)

        # Row 2: Extracted text values
        row_num += 1
        all_field_values = {field: extract_value(plain_text, field) for field in column_headers}
        field_values = [all_field_values[h] for h in filtered_headers]
        for col, value in enumerate(field_values, start=1):
            ws.cell(row=row_num, column=col, value=value)

        # Row 3: Red-colored values
        row_num += 1
        for col, red_text in enumerate(red_texts[:len(filtered_headers)], start=1):
            ws.cell(row=row_num, column=col, value=red_text).font = Font(color="FF0000")

    # Export to Excel
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.success("✅ Excel file created successfully!")
    st.download_button(
        label="📥 Download Excel",
        data=output,
        file_name="output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
